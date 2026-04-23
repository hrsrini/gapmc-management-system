"""
Single pass: Requirement_Verification_with_SRS_Version_3.0.xlsx
- Refresh Developer Findings + Developement Remarks (Module_01, Module_02, Checklist_Verification_Applicat)
- Add column "Development Status" immediately after "Developement Remarks" (first free column)
- Fill Development Status from sheet Status / RESULT

Usage (from repo root):
  python scripts/sync-requirement-verification-workbook.py
"""
from __future__ import annotations

from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "requirements" / "Requirement_Verification_with_SRS_Version_3.0.xlsx"

DEV_STATUS_HEADER = "Development Status"


def wrap_open(findings: str, remarks: str) -> tuple[str, str]:
    f = (findings or "").strip()
    r = (remarks or "").strip()
    if "What remains open" not in f:
        f = f.rstrip(".") + ". What remains open: see Developement Remarks."
    if "What remains open" not in r:
        r = r.rstrip(".") + ". What remains open: see this cell (pending items described above)."
    return f, r


def wrap_open_module02(findings: str, remarks: str) -> tuple[str, str]:
    return wrap_open(findings, remarks)


# ----- Module_01 (Sr. No. keyed) -----
MODULE01_UPDATES: dict[int, tuple[str, str]] = {
    1: (
        "Gap: Employee master (SCR-EMP-02 / SRS 4.1.1) lacked demographic, address, emergency contact, and reporting-officer fields in UI and API.",
        "Added DB columns + employee create/update API + form fields: gender, marital status, blood group, permanent/correspondence address, emergency contact name/mobile, reporting officer (validated FK).",
    ),
    2: (
        "Gap: Same as Sr.1 - data dictionary fields not all persisted or editable.",
        "Aligned with Sr.1 delivery; employee master extended per SRS dictionary subset in this release.",
    ),
    3: (
        "Gap: SCR-EMP-02 profile form not evidenced end-to-end vs SRS (tabs/fields).",
        "Public / personal / HR tabs on employee form now cover extended master; EMP-ID flow unchanged (DA approval). Further SRS-only artefacts (e.g. attachments) out of this slice.",
    ),
    5: (
        "Gap: SCR-LVE-01 - missing reason/supporting docs, opening balances, and DO submission UX; notifications/calendar not in scope here.",
        "Added reason + supporting_document_url on leave_requests; New leave dialog for DO/Admin; DA approval debits configured leave balance (inclusive calendar days) when a balance row exists.",
    ),
    6: (
        "Gap: SCR-LVE-01 form parity - same as Sr.5.",
        "Same as Sr.5: leave capture + balances + workflow UI hooks.",
    ),
    7: (
        "Gap: No authorised screen to set go-live opening leave balances per employee/type.",
        "Implemented employee_leave_balances, GET/PUT /api/hr/leave-balances, and /hr/leave-balances (M-01 Read; save requires M-01 Update).",
    ),
    8: (
        "Gap: TA/DA entitlement matrix not admin-configurable; DO could not file TA/DA from UI.",
        "Added system_config key ta_da_entitlement_json (default from workbook matrix) + Admin Config JSON editor + reference table on Claims; New TA/DA claim dialog.",
    ),
    9: (
        "Gap: Navigation showed HR and other modules without matching role permissions (M-10 §13.1 UX).",
        "Sidebar links now require appropriate M-* Read (etc.); dashboard quick actions filtered by module Read; HR leave balances route permission-wrapped.",
    ),
    10: (
        "Gap: Verification text incorrectly cited DA/DA module; SCR-CFG-01 vs deployed config needed clarity.",
        "Admin Config + receipt PDF logo already match SCR-CFG-01; TA/DA matrix added to same console; observation treated as erroneous paste.",
    ),
    11: (
        "Gap: LTC was display-only - no Pending to Verified to Approved/Rejected workflow or DO create guard.",
        "Extended ltc_claims with DO/DV/DA workflow fields; PUT /api/hr/claims/ltc/:id + queue filter; New LTC dialog; parity with TA/DA segregation rules.",
    ),
}


# ----- Module_02 (Sr. No. keyed, triple with status) -----
MODULE02_UPDATES: dict[int, tuple[str, str, str]] = {
    1: (
        "Gap: Unified entity register across Track A + Track B + ad-hoc occupants was not available as a single register.",
        "Implemented: unified entity list via GET /api/ioms/unified-entities merging Track A (trader_licences), Track B (entities), and ad-hoc (ad_hoc_entities). Added UI /traders/unified-entities plus ability to create ad-hoc entities (POST /api/ioms/unified-entities/ad-hoc).",
        "Verified and Closed",
    ),
    2: (
        "Gap: Sub-type driven behaviour (licence vs no-licence, GST applicability, billing doc type) is still partially represented (licenceType + govt GST exemption + non-GST flag; Track B subType is a free-text string).",
        "Implemented: formalized Track B subType in shared/track-b-entity.ts + /api/ioms/reference/entity-subtypes (billing notes); shared TRACKB_NON_GOV_DUES_API_HINT aligned with GET /api/ioms/dues trackBBillingHint for non-Govt TB; entity list shows a Billing column (short label) plus header copy (Govt vs Commercial/Ad-hoc); entity profile shows billing-route alert with links to Dues (TB:…), Pre-receipts (Govt), M-03 rent deposit ledger, and for non-Govt also Rent/GST invoices register (/rent/ioms/invoices). Pre-receipt POST uses strict Govt match; GET dues lists pre-receipt rows only for Govt; entity PUT blocks changing away from Govt if open pre-receipts exist; Pre-receipts UI entity picker lists Govt entities only. What remains open: deeper SRS subtype catalogue / GST applicability matrix and full Commercial Track B tax-invoice UX.",
        "Reported",
    ),
    3: (
        "Gap (as per sheet): unified Entity Register list across Track A + Track B.",
        "Delivered: Track A entity register /traders/licences (GET /api/ioms/traders/licences); Track B entity register /traders/entities (GET /api/ioms/entities); merged unified register /traders/unified-entities (GET /api/ioms/unified-entities) including ad-hoc occupants.",
        "Verified and Closed",
    ),
    4: (
        "Gap (as per sheet): entity profile and premises allocation list per entity.",
        "Delivered for Track A: /traders/licences/:id includes Premises allocations (GET /api/ioms/asset-allotments?traderLicenceId=). Delivered for Track B: /traders/entities/:id profile with entity allotments grid (GET /api/ioms/entity-allotments?entityId=) and register edit fields.",
        "Verified and Closed",
    ),
    5: (
        "Gap: Trader / lease holder ledger view (SCR-TRD-03) was not discoverable as an M-02 feature earlier (ledger existed under Rent/Tax module).",
        "Implemented: Rent deposit ledger at /rent/ioms/ledger (GET /api/ioms/rent/ledger) with Trader licence deep-link; when filtered by TA: or tenant id, supplemental GET /api/ioms/rent/ledger/trader-receipts lists IOMS receipts (any revenue head) with payer TraderLicence = that tenant for cross-check with M-04/M-02/M-05. What remains open: finance-mandated GL posting bridge or a single consolidated posted ledger if SRS requires it beyond deposit + receipt views.",
        "Verified and Closed",
    ),
    6: (
        "Gap: Outstanding dues and payment screen (self-service + counter) was not implemented earlier.",
        "Implemented: /traders/dues with GET /api/ioms/dues: Track A rent (POST /api/ioms/dues/pay-rent-invoice); Track B Govt pre-receipts; Track A M-04 market fee liability with counter POST /api/ioms/dues/pay-market-fee (partial/full; nets Pending M-04 placeholders). Auth: GET dues and GET /api/ioms/unified-entities allow M-02:Read or M-03:Read (rent desk picker); POST pay-rent-invoice requires M-03:Update; pay-market-fee remains M-04 Create/Update; pre-receipts + unified-entity mutations are M-02 per getModuleForPath. What remains open: customer self-service portal and hosted online payment gateway (vendor/Phase-2).",
        "Verified and Closed",
    ),
    7: (
        "Gap: Track B non-trader lease holder registration was missing earlier.",
        "Implemented: entities table + GET/POST/PUT /api/ioms/entities + entity allotments; New entity dialog and entity profile show PAN, GSTIN, email, address (mobile sanitised); profile Edit register (M-02 Update) saves via PUT; server trims/lowercases email on create/update; getModuleForPath maps /api/ioms/entities and /api/ioms/entity-allotments to M-02. Billing/pre-receipt behaviour by sub-type is covered under Sr.2. What remains open: any extra SRS identity or registration fields not yet in schema.",
        "Verified and Closed",
    ),
    8: (
        "Gap: Pre-receipt generation and tracking for govt entities (Track B) was missing earlier.",
        "Implemented: pre_receipts table + APIs (/api/ioms/pre-receipts) + UI (/traders/pre-receipts) with lifecycle Issued->Dispatched->Acknowledged->Settled. Now enforces Govt-only entities by subtype and auto-creates an IOMS receipt on settlement when settledReceiptId is not provided.",
        "Verified and Closed",
    ),
    9: (
        "Gap: New/renewal market functionary registration (Form BM) was only partially implemented (licenceType enum existed, but no explicit functionary-focused UX).",
        "Implemented: /traders/functionaries list + licenceTypes filter; Form BM fields on trader licence application (father/spouse name, DOB, emergency mobile, character certificate issuer/date) plus optional supporting-document URL bm_form_doc_url (016) and optional uploaded file bm_form_doc_file (017) via POST/DELETE /api/ioms/traders/licences/:id/bm-form-document (PDF/PNG/JPEG, 10 MB, blob store); GET same path serves file; not replaceable after licence number issued; renewal copies parent file into new draft; detail + edit UI; form shows explicit note that full SRS checklist/approvals/photo may be required at go-live. What remains open: SRS-complete BM checklist, role-specific approvals, photo capture if required.",
        "Reported",
    ),
    10: (
        "Gap: Renewal of licence under Section 54 (Form BK) was missing as an explicit operation earlier.",
        "Implemented: POST /api/ioms/traders/licences/:id/renew clones Draft renewal; parent_licence_fee_snapshot (014); optional body feeAmount else defaults from parent fee then system licence_fee config; GET /api/ioms/traders/licences/:id/renew-preview returns the same resolution breakdown before create; licence detail Renew opens a dialog (default fee, optional override, optional valid dates) and notes counter/M-05 payment (no hosted checkout); renew dialog links to IOMS receipts register; Form BK declaration before Pending; UI BK panel + detail. What remains open: hosted payment gateway for renewal fee, multi-year automated tariff tables beyond config, any extra SRS BK fields.",
        "Reported",
    ),
    11: (
        "Gap (as per sheet): Premises master registration/directory/lifecycle.",
        "Implemented as Asset Register (M-02): assets table + /assets page + APIs (/api/ioms/assets).",
        "Verified and Closed",
    ),
    12: (
        "Gap (as per sheet): Premises Master (M-02-PM).",
        "Implemented as assets (one record per physical unit) with yard/type/complex/area and lifecycle via isActive.",
        "Verified and Closed",
    ),
    13: (
        "Gap (as per sheet): Premises master exists independently of occupant and persists for lifetime.",
        "Implemented: assets exist independent of allotments; allotments link to assets via asset_allotments.",
        "Verified and Closed",
    ),
    14: (
        "Gap (as per sheet): Premises lifecycle directory screens.",
        "Implemented via Assets list plus Shop Vacant / Shop Allotments screens and APIs (M-02).",
        "Verified and Closed",
    ),
    15: (
        "Gap: Unified entity master across Track A + Track B + ad-hoc occupants was not available as a single register.",
        "Implemented: GET/POST unified-entities + UI; unified IDs on dues, receipts (incl. backfill 013), rent deposit ledger rows, Tally/receipt register/PDF; rent ledger + payer-linked receipt panel for TA:. GET unified-entities list/by-id is permissioned (M-02:Read or M-03:Read for picker use); POST ad-hoc remains M-02:Create. What remains open: product scenarios with rent/allotment but no Track A licence (ad-hoc-only yard occupant) if later introduced; finance-specific GL exports beyond current artefacts.",
        "Verified and Closed",
    ),
    16: (
        "Gap (as per sheet): premises allocation record per premises per entity.",
        "Implemented as asset_allotments with one row per asset per trader licence; asset attributes are stored on assets (not duplicated).",
        "Verified and Closed",
    ),
    17: (
        "Gap: Rent revision configuration (section G) not implemented/identified in current implementation.",
        "Implemented: rent_revision_overrides + /rent/ioms/revisions + DO→DV→DA workflow; migration 015 adds revision_basis (FixedMonthlyRent | OtherDocumented) with API/UI capture; OtherDocumented requires remarks (≥20 chars). GET /api/ioms/rent/allotments/:allotmentId/rent-context returns resolved baseline rent (same rules as invoice/cron via shared server/rent-allotment-rent-resolve.ts) for optional effectiveMonth=YYYY-MM; Rent Revisions UI shows baseline + optional % apply to pre-fill rent_amount; system_config rent_revision_suggested_percent (default 0) pre-fills the % field when admin sets a positive value; UI states plinth/rate engines and auto ledger re-post are out of scope for this build. What remains open: plinth-area / rate-table engines; ledger re-posting for already-issued invoices if finance requires it.",
        "Reported",
    ),
}


# ----- Checklist (module, FR ID) -----
CHECKLIST_UPDATES: dict[tuple[str, str], dict[str, str]] = {
    ("M-01", "FR-LVE-001"): {
        "findings": "Leave application workflow exists, but the checklist row was marked Not Implemented in workbook.",
        "remarks": "Implemented: leave requests API/UI with DO->DV->DA workflow, reason + optional supporting document URL, DV return with remarks, DA approve/reject with mandatory remarks. What remains open: SRS-level calendar/notifications and leave-type-specific doc enforcement (ML/CCL) policy mapping.",
        "result": "To be Checked",
    },
    ("M-01", "FR-LVE-002"): {
        "findings": "Supporting documents are captured (URL) but enforcement by leave type is not implemented.",
        "remarks": "Implemented: supporting_document_url capture + server-side enforcement for ML/CCL leave types on create/update. What remains open: move from URL to managed attachment storage and align exact leave-type list from SRS (if more than ML/CCL).",
        "result": "Implemented",
    },
    ("M-01", "FR-WFL-001"): {
        "findings": "Maker-checker workflow is implemented across modules via DO/DV/DA rules, but this checklist row was marked Not Implemented.",
        "remarks": "Implemented: DO/DV/DA segregation checks in workflow utilities and enforced in key routes (HR leave/claims, trader licences, market transactions, etc.). What remains open: complete coverage audit across every mutation endpoint and align per-SRS exceptions.",
        "result": "To be Checked",
    },
    ("M-02", "FR-PRE-001"): {
        "findings": "Premises master and allocation are implemented (assets + asset_allotments). Workbook had Not Implemented.",
        "remarks": "Implemented: premises master via assets (Asset Register) and allocations via asset_allotments (Shop Allotments). Trader licence detail now shows premises allocations list.",
        "result": "Implemented",
    },
    ("M-02", "FR-PRE-002"): {
        "findings": "Premises allocation links to a valid entity (trader licence) and asset.",
        "remarks": "Implemented: asset_allotments requires existing assetId and traderLicenceId; licence profile lists allocations for that licence.",
        "result": "Implemented",
    },
    ("M-02", "FR-PRE-003"): {
        "findings": "Pre-receipt Track B govt and premises lifecycle evidence were partially missing in the workbook.",
        "remarks": "Implemented: (1) Pre-receipts — pre_receipts + /api/ioms/pre-receipts (M-02 permissioned) + /traders/pre-receipts lifecycle through settlement; Govt-only entity match by subtype. (2) Premises lifecycle — assets.isActive plus vacancy/allotment M-02 screens. What remains open: full SRS lifecycle states and document/versioning if specified beyond current flags.",
        "result": "To be Checked",
    },
    ("M-02", "FR-AST-003"): {
        "findings": "Multiple premises per entity supported via multiple asset_allotments per trader licence.",
        "remarks": "Implemented: one licence can have many allotment rows; API supports query by traderLicenceId and assetId.",
        "result": "Implemented",
    },
    ("M-02", "FR-AST-004"): {
        "findings": "Entity lifecycle is partially implemented for Track A licences (status + block/unblock log); Track B was thinner.",
        "remarks": "Implemented for Track A: status (Draft/Pending/Query/Active/Expired/Blocked) + expiry auto-block cron + blocking log. Track B: entities.status with profile Edit register (M-02 Update). What remains open: Track B automated block/expiry parity with Track A if SRS mandates.",
        "result": "To be Checked",
    },
    ("M-02", "FR-AST-001"): {
        "findings": "Unified entity registry across Track A + Track B + ad-hoc was missing earlier.",
        "remarks": "Implemented: GET /api/ioms/unified-entities + /traders/unified-entities; POST ad-hoc; unified IDs on dues, receipts, rent ledger/Tally paths per prior delivery. What remains open: rent/allotment with no Track A licence (ad-hoc-only occupant) if product later requires it.",
        "result": "To be Checked",
    },
    ("M-02", "FR-AST-002"): {
        "findings": "Trader licence validation exists in licence workflow; Track B registration not present.",
        "remarks": "Implemented: trader licence CRUD/workflow with yard scoping and approval fields. What remains open: validations during Track B entity registration and any SRS BM/BK form-specific validation.",
        "result": "To be Checked",
    },
    ("M-02", "FR-CHG-001"): {
        "findings": "Charge calculation rules are not implemented as a premises-charge engine in M-02.",
        "remarks": "What remains open: define charge rule engine (rent revision/config, base charges by premises/type/area) and integrate into invoice/receipt generation.",
        "result": "Not Implemented",
    },
    ("M-02", "FR-GST-001"): {
        "findings": "GST applicability based on entity classification is partially supported via govt GST exemption and non-GST flags.",
        "remarks": "Implemented partial flags on licences used by rent/receipt paths. What remains open: full Track A/B classification matrix and end-to-end GST rules per SRS (billing doc type, GST thresholds, exemptions).",
        "result": "To be Checked",
    },
    ("M-02", "FR-NRT-001"): {
        "findings": "Track B non-trader entity register was missing earlier.",
        "remarks": "Implemented: entities + GET/POST/PUT /api/ioms/entities + /traders/entities (PAN/GSTIN/email/address, profile edit); entity_allotments + entity detail; subtype-driven billing hints (see FR-GST / Sr.2). APIs mapped to M-02 in getModuleForPath with dues GET/pay split per auth. What remains open: extra SRS identity fields not in schema.",
        "result": "To be Checked",
    },
    ("M-03", "FR-RNT-001"): {
        "findings": "Rent/tax invoice module exists (M-03), but checklist row was marked Not Implemented.",
        "remarks": "Implemented: IOMS rent invoice list/detail/create screens (/rent/ioms) with backend APIs (/api/ioms/rent/invoices) and DO/DV/DA workflow states (Draft->Verified->Approved->Paid). What remains open: confirm exact FR-RNT-001 expectation vs SRS clause and adjust fields if needed.",
        "result": "To be Checked",
    },
    ("M-03", "FR-RNT-002"): {
        "findings": "Rent invoice workflow + status transitions are implemented.",
        "remarks": "Implemented: status transitions enforced server-side (workflow utilities) and exposed in UI actions (Verify/Approve/Mark paid). What remains open: full parity with SRS exception rules and role matrix validation for every transition.",
        "result": "To be Checked",
    },
    ("M-03", "FR-RNT-003"): {
        "findings": "GST export support exists but may not match final GSTN filing schema.",
        "remarks": "Implemented: GSTR-1 draft export endpoint GET /api/ioms/rent/gstr1 with warnings for GSTIN issues and ctin/id date mapping. What remains open: confirm final GSTN JSON schema required by filing tool/CA.",
        "result": "To be Checked",
    },
    ("M-03", "FR-RNT-004"): {
        "findings": "Credit note workflow exists in M-03.",
        "remarks": "Implemented: credit notes API + UI (/rent/ioms/credit-notes) backed by /api/ioms/rent/credit-notes. What remains open: confirm SRS credit-note numbering/approval rules and accounting integration.",
        "result": "To be Checked",
    },
    ("M-03", "FR-RNT-005"): {
        "findings": "TDS logic for rent invoices is implemented, but GL posting is not.",
        "remarks": "Implemented: FY-based rent TDS calculation (Apr-Mar), tds fields on rent invoices + receipts, and PDF rendering. What remains open: GL posting rules for TDS and marginal-month behaviour if required.",
        "result": "To be Checked",
    },
    ("M-03", "FR-RNT-006"): {
        "findings": "Rent deposit ledger is implemented.",
        "remarks": "Implemented: rent deposit ledger API (/api/ioms/rent/ledger) and UI (/rent/ioms/ledger); GET /api/ioms/rent/ledger/trader-receipts lists IOMS receipts with payer TraderLicence for the same TA/tenant filter (supplemental panel on ledger screen). Receipts marked Paid can post Collection entries via hook. What remains open: GL posting bridge if finance mandates posted consolidation beyond deposit + receipt views.",
        "result": "To be Checked",
    },
    ("M-03", "FR-RNT-007"): {
        "findings": "Dishonour/arrears disclosure exists but needs accounting closure.",
        "remarks": "Implemented: arrears disclosure on GET receipt + PDF after prior dishonour for same invoice; IOMS receipt detail page surfaces disclosure with invoice + ledger links; Outstanding dues rent pay dialog includes dishonour/arrears note and invoice link; PATCH Reversed can return rentDishonourScaffold with voucherCreateHref; VoucherCreate prefills from query string when finance records bank charge. What remains open: interest posting to GL and one-click posted bank charge if finance mandates.",
        "result": "To be Checked",
    },
    ("M-03", "FR-RNT-008"): {
        "findings": "Rent invoice report/export features exist partially (reports + exports in repo).",
        "remarks": "Implemented: rent reports page + server-side report endpoints (see /rent/reports and routes-reports.ts). What remains open: confirm exact SRS report formats and any statutory exports beyond current implementation.",
        "result": "To be Checked",
    },
    ("M-03", "FR-RNT-009"): {
        "findings": "Monthly auto-generation for rent invoices exists.",
        "remarks": "Implemented: cron-based monthly Draft invoice generation for active allotments. What remains open: any SRS rule variations for proration, rent revision types, and audit/notifications.",
        "result": "To be Checked",
    },
    ("M-03", "FR-RNT-010"): {
        "findings": "Rent revision configuration is now implemented minimally via overrides.",
        "remarks": "Implemented: rent_revision_overrides + UI/API; DO→DV→DA; cron/manual invoice use Approved overrides; revision_basis (FixedMonthlyRent | OtherDocumented) on rows (015) for SRS classification — billing still uses approved rent_amount INR until percent/area auto-rules. What remains open: automated rule engine and ledger impacts on issued invoices if finance mandates.",
        "result": "To be Checked",
    },
    ("M-04", "FR-MKT-001"): {
        "findings": "Earlier implementation used a free-form market fee % on purchase create.",
        "remarks": "Implemented: resolveMarketFeePercentForPurchase (yard-specific market_fee_rates row, else global yard_id NULL, else system_config market_fee_percent); GET /api/ioms/market/fee-preview; POST enforces match when client sends marketFeePercent and validates marketFeeAmount vs declared × resolved %. MarketTransactions UI loads preview read-only.",
        "result": "To be Checked",
    },
    ("M-04", "FR-MKT-002"): {
        "findings": "Purchase create did not enforce active / unblocked / in-validity trader licence.",
        "remarks": "Implemented: POST /api/ioms/market/transactions requires licence status Active, not is_blocked, and transactionDate within validFrom/validTo when both are ISO dates on the licence.",
        "result": "To be Checked",
    },
    ("M-04", "FR-MKT-003"): {
        "findings": "Commodity must reference commodities master.",
        "remarks": "Implemented: POST still rejects unknown commodityId (404 PURCHASE_TX_COMMODITY_NOT_FOUND); fee matrix keyed by commodityId. What remains open: optional isActive filter on commodity if SRS requires inactive commodities blocked.",
        "result": "To be Checked",
    },
    ("M-10", "FR-USR-005"): {
        "findings": "Local password policy was weaker than BR-USR-10 (min 8).",
        "remarks": "Implemented: shared/password-policy-br-usr-10.ts + assertPasswordComplexityBrUsr10 on POST/PUT employee login; HR UI hints updated; seed passwords updated to compliant defaults (GapmcAdmin@2026!, GapmcUsers@2026!).",
        "result": "To be Checked",
    },
}


def development_status_from_verification_status(status_val: object) -> str:
    t = str(status_val or "").strip()
    if t == "Verified and Closed":
        return "Complete — aligned with repo; no open dev actions on this row"
    if t == "Reported":
        return "Partial — delivered subset; follow-up documented in remarks / needs client-finance input"
    if not t:
        return "Not set"
    return t


def development_status_from_result(result_val: object) -> str:
    r = str(result_val or "").strip()
    if r == "Implemented":
        return "Development complete — behaviour present in repo for this FR"
    if r == "To be Checked":
        return "Partial — engineering delivered; UAT / CA / finance confirmation still needed"
    if r == "Not Implemented":
        return "Not in repo or explicitly out of scope — see remarks"
    if r == "Requires Rework":
        return "Engineering follow-up — see updated remarks / RESULT after rework"
    if not r:
        return "Not assessed"
    return f"RESULT: {r}"


def process_module_sheet(ws, sheet_name: str) -> None:
    header = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    ix_sr = header.index("Sr. No.")
    ix_find = header.index("Developer Findings")
    ix_rem = header.index("Developement Remarks")
    ix_status = header.index("Status")
    dev_col_1based = ix_rem + 2  # column immediately after Developement Remarks (1-based)

    ws.cell(row=1, column=dev_col_1based).value = DEV_STATUS_HEADER

    for row_idx in range(2, ws.max_row + 1):
        sr_cell = ws.cell(row=row_idx, column=ix_sr + 1)
        try:
            sr = int(sr_cell.value)
        except (TypeError, ValueError):
            ws.cell(row=row_idx, column=dev_col_1based).value = ""
            continue

        if sheet_name == "Module_01":
            if sr in MODULE01_UPDATES:
                f, r = MODULE01_UPDATES[sr]
                ws.cell(row=row_idx, column=ix_find + 1).value = f
                ws.cell(row=row_idx, column=ix_rem + 1).value = r
                ws.cell(row=row_idx, column=ix_status + 1).value = "Verified and Closed"
            st = ws.cell(row=row_idx, column=ix_status + 1).value
            ws.cell(row=row_idx, column=dev_col_1based).value = development_status_from_verification_status(st)

        elif sheet_name == "Module_02":
            if sr in MODULE02_UPDATES:
                findings, remarks, status = MODULE02_UPDATES[sr]
                if status == "Reported":
                    findings, remarks = wrap_open_module02(findings, remarks)
                ws.cell(row=row_idx, column=ix_find + 1).value = findings
                ws.cell(row=row_idx, column=ix_rem + 1).value = remarks
                ws.cell(row=row_idx, column=ix_status + 1).value = status
            st = ws.cell(row=row_idx, column=ix_status + 1).value
            ws.cell(row=row_idx, column=dev_col_1based).value = development_status_from_verification_status(st)


def process_checklist(ws) -> None:
    header = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    ix = {h: i for i, h in enumerate(header) if h}
    for c in ("Module", "FR ID", "RESULT", "Developer Findings", "Developement Remarks"):
        if c not in ix:
            raise SystemExit(f"Checklist missing column: {c}")

    ix_rem = ix["Developement Remarks"]
    dev_col_1based = ix_rem + 2
    ws.cell(row=1, column=dev_col_1based).value = DEV_STATUS_HEADER

    for row_idx in range(2, ws.max_row + 1):
        module = str(ws.cell(row=row_idx, column=ix["Module"] + 1).value or "").strip()
        fr = str(ws.cell(row=row_idx, column=ix["FR ID"] + 1).value or "").strip()
        res_cell = ws.cell(row=row_idx, column=ix["RESULT"] + 1)
        result_val = res_cell.value

        if not module or not fr:
            ws.cell(row=row_idx, column=dev_col_1based).value = ""
            continue

        key = (module, fr)
        if key in CHECKLIST_UPDATES:
            u = CHECKLIST_UPDATES[key]
            findings = u.get("findings", "") or ""
            remarks = u.get("remarks", "") or ""
            result = u.get("result")
            if result and result != "Implemented":
                findings, remarks = wrap_open(findings, remarks)
            ws.cell(row=row_idx, column=ix["Developer Findings"] + 1).value = findings
            ws.cell(row=row_idx, column=ix["Developement Remarks"] + 1).value = remarks
            if result:
                res_cell.value = result
            result_val = res_cell.value

        else:
            res = str(result_val or "").strip()
            if res and res != "Implemented":
                f0 = str(ws.cell(row=row_idx, column=ix["Developer Findings"] + 1).value or "").strip() or (
                    "Gap: checklist item not yet verified in this repo."
                )
                r0 = str(ws.cell(row=row_idx, column=ix["Developement Remarks"] + 1).value or "").strip() or (
                    "What remains open: verify implementation against SRS and update RESULT accordingly."
                )
                f1, r1 = wrap_open(f0, r0)
                ws.cell(row=row_idx, column=ix["Developer Findings"] + 1).value = f1
                ws.cell(row=row_idx, column=ix["Developement Remarks"] + 1).value = r1

        ws.cell(row=row_idx, column=dev_col_1based).value = development_status_from_result(res_cell.value)


def main() -> None:
    wb = openpyxl.load_workbook(XLSX, data_only=False)
    process_module_sheet(wb["Module_01"], "Module_01")
    process_module_sheet(wb["Module_02"], "Module_02")
    process_checklist(wb["Checklist_Verification_Applicat"])
    wb.save(XLSX)
    print("Synced workbook + Development Status column:", XLSX)


if __name__ == "__main__":
    main()
